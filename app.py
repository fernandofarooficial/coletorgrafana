import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import csv
import os
import re
import mysql.connector
from mysql.connector import Error
from flask import Flask, render_template, request, send_file
import threading
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

app = Flask(__name__)

URL = "https://zions.grafana.net/public-dashboards/f8bfe8e54a494de9b59218743d70381c"
CSV_FILE = "dados_grafana.csv"

DB_CONFIG = {
    'user': os.environ.get('DB_USER', 'fefa_dev'),
    'password': os.environ.get('DB_PASSWORD', 'Fd7493dt'),
    'host': os.environ.get('DB_HOST', '72.60.58.241'),
    'port': int(os.environ.get('DB_PORT', '3306')),
    'database': os.environ.get('DB_NAME', 'bufunfa'),
    'time_zone': os.environ.get('DB_TIMEZONE', '-03:00'),
    'autocommit': True,
    'connect_timeout': 30,
    'connection_timeout': 30
}


def obter_horario_brasil():
    """Retorna o horário atual no timezone do Brasil (UTC-3)"""
    # Timezone do Brasil (São Paulo)
    tz_brasil = pytz.timezone('America/Sao_Paulo')
    # Obtém o horário atual UTC e converte para o timezone do Brasil
    return datetime.now(pytz.utc).astimezone(tz_brasil)


def conectar_banco():
    """Conecta ao banco de dados MySQL"""
    try:
        conexao = mysql.connector.connect(**DB_CONFIG)
        if conexao.is_connected():
            return conexao
    except Error as e:
        print(f"  Erro ao conectar ao banco de dados: {e}")
        return None


def extrair_dados_com_scroll_profundo(driver):
    """Extrai dados com scroll PROFUNDO para carregar TODOS os Prompts"""
    try:
        print("  Carregando página inteira (scroll profundo)...")
        time.sleep(3)

        # Scroll até o topo
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(1)

        # Scroll INCREMENTAL para carregar tudo
        ultimo_height = driver.execute_script("return document.body.scrollHeight")
        scrolls = 0

        while scrolls < 20:  # Força mínimo 20 scrolls
            driver.execute_script("window.scrollBy(0, 1000);")
            time.sleep(0.5)

            nova_height = driver.execute_script("return document.body.scrollHeight")
            scrolls += 1

            if nova_height == ultimo_height:
                # Se altura não mudou, tenta mais alguns scrolls
                if scrolls > 10:
                    break
            ultimo_height = nova_height

        # Volta para cima
        driver.execute_script("window.scrollTo(0, 0);")
        time.sleep(3)

        # Extrai todo o texto
        texto_pagina = driver.execute_script("return document.body.innerText;")

        print(f"  Página carregada ({len(texto_pagina)} caracteres)")
        print(f"  Scrolls realizados: {scrolls}")

        # Debug: mostra quantos "Prompt " encontrou
        num_prompts = texto_pagina.count("Prompt ")
        print(f"  Prompts encontrados no texto: {num_prompts}")

        # Procura pelos Prompts
        prompts_dados = {}

        # Divide o texto por "Prompt "
        partes = texto_pagina.split("Prompt ")

        print(f"  Processando {len(partes) - 1} seções...")

        for idx, parte in enumerate(partes[1:], start=1):
            print(f"  Processando Prompt {idx}...")

            dados_prompt = {}

            # Padrões para extrair valores (mais robusto)
            padroes = {
                'CPU': r'CPU\s+([\d.]+)\s*%',
                'RAM': r'RAM\s+([\d.]+)\s*%',
                'GPU': r'GPU\s+([\d.]+)\s*%',
                'HD': r'HD\s+([\d.]+)\s*%',
                'SWAP': r'SWAP\s+([\d.]+)\s*%',
                'Infer time': r'Infer time\s+([\d.]+)',
                'CPU Temperature': r'CPU Temperature\s+([\d.]+)',
                'GPU Temperature': r'Temperatura GPU\s+([\d.]+)',
                'TCP Latency GOOG': r'goog\s+([\d.]+)',
                'TCP Latency MON': r'mon\s+([\d.]+)',
                'Câmeras On': r'Câmeras On\s+([\d]+)',
                'Câmeras Inativas': r'Câmeras Inativas\s+([\d]+)',
                'Câmeras Ociosas': r'Câmeras Ociosas\s+([\d]+)',
                'Total Câmeras': r'Total Câmeras\s+([\d]+)'
            }

            campos_encontrados = 0

            for campo, padrao in padroes.items():
                match = re.search(padrao, parte, re.IGNORECASE)
                if match:
                    try:
                        valor = float(match.group(1))
                        dados_prompt[campo] = valor
                        print(f"    {campo}: {valor}")
                        campos_encontrados += 1
                    except:
                        dados_prompt[campo] = None
                else:
                    dados_prompt[campo] = None

            # Se encontrou pelo menos alguns campos, salva este Prompt
            if campos_encontrados > 0:
                prompts_dados[f'Prompt {idx}'] = dados_prompt
                print(f"    Total de campos encontrados: {campos_encontrados}")
            else:
                print(f"    ⚠️  Nenhum campo encontrado para Prompt {idx}")

        print(f"  Total de Prompts com dados: {len(prompts_dados)}")

        return prompts_dados if prompts_dados else None

    except Exception as e:
        print(f"  Erro: {e}")
        import traceback
        traceback.print_exc()
        return None


def extrair_dados_selenium():
    """Abre Selenium e extrai dados"""
    driver = None
    try:
        horario_brasil = obter_horario_brasil()
        print(f"[{horario_brasil.strftime('%d/%m/%Y %H:%M:%S')} Brasil] Iniciando coleta...")
        print(f"  URL: {URL}")

        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-software-rasterizer")
        options.add_argument("--window-size=1920,1200")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

        print("  Inicializando Chrome...")
        driver = webdriver.Chrome(options=options)
        print("  Chrome inicializado com sucesso")

        print(f"  Acessando URL: {URL}")
        driver.get(URL)
        print("  URL acessada com sucesso")

        print("  Aguardando carregamento da página...")
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.TAG_NAME, "span"))
        )
        print("  Página carregada")

        dados = extrair_dados_com_scroll_profundo(driver)
        driver.quit()
        print("  Chrome fechado")

        return dados

    except Exception as e:
        print(f"  ❌ ERRO no Selenium: {e}")
        print(f"  Tipo do erro: {type(e).__name__}")
        import traceback
        print("  Stack trace completo:")
        traceback.print_exc()

        try:
            if driver:
                driver.quit()
                print("  Chrome fechado após erro")
        except:
            pass
        return None


def salvar_em_csv(dados_dict):
    """Salva dados em CSV"""
    if not dados_dict:
        print("  Sem dados para salvar")
        return False

    arquivo_existe = os.path.isfile(CSV_FILE)

    try:
        with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
            campos = ['Data/Hora', 'Prompt', 'CPU', 'RAM', 'GPU', 'HD', 'SWAP',
                      'Infer time', 'CPU Temperature', 'GPU Temperature',
                      'TCP Latency GOOG', 'TCP Latency MON',
                      'Câmeras On', 'Câmeras Inativas', 'Câmeras Ociosas', 'Total Câmeras']
            writer = csv.DictWriter(f, fieldnames=campos)

            if not arquivo_existe:
                writer.writeheader()

            timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

            # Itera por TODOS os prompts encontrados
            for prompt_key in sorted(dados_dict.keys()):
                row = {'Data/Hora': timestamp, 'Prompt': prompt_key}

                for campo in campos[2:]:
                    row[campo] = dados_dict[prompt_key].get(campo, None)

                writer.writerow(row)

            # Print resumo
            print("  Dados salvos:")
            for prompt_key in sorted(dados_dict.keys()):
                cpu = dados_dict[prompt_key].get('CPU', 'N/A')
                ram = dados_dict[prompt_key].get('RAM', 'N/A')
                gpu = dados_dict[prompt_key].get('GPU', 'N/A')
                hd = dados_dict[prompt_key].get('HD', 'N/A')
                print(f"  ✓ {prompt_key}: CPU={cpu}% RAM={ram}% GPU={gpu}% HD={hd}%")

        return True

    except Exception as e:
        print(f"  Erro ao salvar CSV: {e}")
        return False


def salvar_em_banco(dados_dict):
    """Salva dados no banco de dados MySQL"""
    if not dados_dict:
        print("  Sem dados para salvar")
        return False

    # Mapeamento de campos CSV para sufixos da tabela
    mapeamento_campos = {
        'CPU': 'cpu',
        'RAM': 'ram',
        'GPU': 'gpu',
        'HD': 'hd',
        'SWAP': 'swap',
        'Infer time': 'infer_time',
        'CPU Temperature': 'cpu_temp',
        'GPU Temperature': 'gpu_temp',
        'TCP Latency GOOG': 'tcp_latency',
        'TCP Latency MON': 'mon_latency',
        'Câmeras On': 'cam_on',
        'Câmeras Inativas': 'cam_off',
        'Câmeras Ociosas': 'cam_idle',
        'Total Câmeras': 'total_cam'
    }

    try:
        conexao = conectar_banco()
        if not conexao:
            print("  Erro: Não foi possível conectar ao banco de dados")
            return False

        cursor = conexao.cursor()

        # Prepara os dados para uma única linha com horário do Brasil
        timestamp_brasil = obter_horario_brasil()
        # Remove timezone info para compatibilidade com MySQL
        timestamp = timestamp_brasil.replace(tzinfo=None)
        dados_linha = {'quando': timestamp}

        print(f"  Horário da coleta (Brasil): {timestamp_brasil.strftime('%d/%m/%Y %H:%M:%S')}")

        # Processa cada Prompt (Prompt 1, Prompt 2, Prompt 3)
        for prompt_num in range(1, 4):
            prompt_key = f'Prompt {prompt_num}'
            prefixo = f'p{prompt_num}_'

            if prompt_key in dados_dict:
                dados_prompt = dados_dict[prompt_key]
                for campo_csv, sufixo in mapeamento_campos.items():
                    campo_tabela = f'{prefixo}{sufixo}'
                    valor = dados_prompt.get(campo_csv)
                    dados_linha[campo_tabela] = valor
            else:
                # Se o Prompt não foi encontrado, preenche com None
                for sufixo in mapeamento_campos.values():
                    campo_tabela = f'{prefixo}{sufixo}'
                    dados_linha[campo_tabela] = None

        # Monta o SQL INSERT
        colunas = ', '.join(dados_linha.keys())
        placeholders = ', '.join(['%s'] * len(dados_linha))
        sql = f"INSERT INTO grafana ({colunas}) VALUES ({placeholders})"

        # Executa o INSERT
        cursor.execute(sql, list(dados_linha.values()))
        conexao.commit()

        # Print resumo
        print("  Dados salvos no banco de dados:")
        for prompt_num in range(1, 4):
            prompt_key = f'Prompt {prompt_num}'
            if prompt_key in dados_dict:
                cpu = dados_dict[prompt_key].get('CPU', 'N/A')
                ram = dados_dict[prompt_key].get('RAM', 'N/A')
                gpu = dados_dict[prompt_key].get('GPU', 'N/A')
                hd = dados_dict[prompt_key].get('HD', 'N/A')
                print(f"  ✓ {prompt_key}: CPU={cpu}% RAM={ram}% GPU={gpu}% HD={hd}%")
            else:
                print(f"  ✗ {prompt_key}: Sem dados")

        cursor.close()
        conexao.close()
        return True

    except Error as e:
        print(f"  Erro ao salvar no banco de dados: {e}")
        if conexao and conexao.is_connected():
            conexao.close()
        return False


def buscar_ultimas_linhas(limite=50):
    """Busca as últimas N linhas do banco de dados"""
    try:
        conexao = conectar_banco()
        if not conexao:
            return []

        cursor = conexao.cursor(dictionary=True)

        # Campos que serão exibidos
        campos = [
            'quando',
            'p1_cpu', 'p1_ram', 'p1_hd', 'p1_gpu', 'p1_swap', 'p1_infer_time', 'p1_gpu_temp', 'p1_total_cam', 'p1_cam_on', 'p1_cam_off', 'p1_cam_idle',
            'p2_cpu', 'p2_ram', 'p2_hd', 'p2_gpu', 'p2_swap', 'p2_infer_time', 'p2_gpu_temp', 'p2_total_cam', 'p2_cam_on', 'p2_cam_off', 'p2_cam_idle',
            'p3_cpu', 'p3_ram', 'p3_hd', 'p3_gpu', 'p3_swap', 'p3_infer_time', 'p3_gpu_temp', 'p3_total_cam', 'p3_cam_on', 'p3_cam_off', 'p3_cam_idle'
        ]

        sql = f"SELECT {', '.join(campos)} FROM grafana ORDER BY quando DESC LIMIT %s"
        cursor.execute(sql, (limite,))

        resultados = cursor.fetchall()

        cursor.close()
        conexao.close()

        return resultados

    except Error as e:
        print(f"  Erro ao buscar dados: {e}")
        return []


@app.route('/')
def index():
    """Página principal que exibe os dados"""
    dados = buscar_ultimas_linhas(50)
    horario_atualizacao = obter_horario_brasil()
    return render_template('index.html', dados=dados, horario_atualizacao=horario_atualizacao)


@app.route('/status')
def status():
    """Rota de status para verificar se a aplicação está funcionando"""
    try:
        conexao = conectar_banco()
        db_status = "conectado" if conexao else "desconectado"
        if conexao:
            conexao.close()

        horario_brasil = obter_horario_brasil()
        return {
            "status": "online",
            "database": db_status,
            "thread_coleta": "ativa" if thread_coleta.is_alive() else "inativa",
            "timestamp_utc": datetime.now().isoformat(),
            "timestamp_brasil": horario_brasil.isoformat(),
            "timezone": "America/Sao_Paulo (UTC-3)"
        }
    except Exception as e:
        return {"status": "erro", "mensagem": str(e)}, 500


@app.route('/exportar_excel')
def exportar_excel():
    """Exporta dados para Excel com três abas (uma para cada servidor)"""
    try:
        # Obtém quantidade de registros (default 50)
        quantidade = request.args.get('quantidade', 50, type=int)

        # Busca os dados
        dados = buscar_ultimas_linhas(quantidade)

        if not dados:
            return "Nenhum dado disponível para exportar", 404

        # Cria o workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove a planilha padrão

        # Estilo para cabeçalho
        header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Estilo para células
        cell_alignment = Alignment(horizontal="center", vertical="center")

        # Define os cabeçalhos
        cabecalhos = ['Data/Hora', 'CPU (%)', 'RAM (%)', 'HD (%)', 'GPU (%)', 'SWAP (%)',
                      'Infer Time', 'GPU Temp', 'Total Câm', 'Câm On', 'Câm Off', 'Câm Idle']

        # Cria três abas (uma para cada servidor)
        for prompt_num in range(1, 4):
            # Cria a aba
            ws = wb.create_sheet(title=f"Prompt {prompt_num}")

            # Adiciona os cabeçalhos
            ws.append(cabecalhos)

            # Estiliza o cabeçalho
            for col_num, _ in enumerate(cabecalhos, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment

            # Adiciona os dados
            for linha in dados:
                row_data = [
                    linha['quando'].strftime('%d/%m/%Y %H:%M:%S') if linha['quando'] else 'N/A',
                    round(linha[f'p{prompt_num}_cpu'], 1) if linha[f'p{prompt_num}_cpu'] is not None else '-',
                    round(linha[f'p{prompt_num}_ram'], 1) if linha[f'p{prompt_num}_ram'] is not None else '-',
                    round(linha[f'p{prompt_num}_hd'], 1) if linha[f'p{prompt_num}_hd'] is not None else '-',
                    round(linha[f'p{prompt_num}_gpu'], 1) if linha[f'p{prompt_num}_gpu'] is not None else '-',
                    round(linha[f'p{prompt_num}_swap'], 1) if linha[f'p{prompt_num}_swap'] is not None else '-',
                    int(linha[f'p{prompt_num}_infer_time']) if linha[f'p{prompt_num}_infer_time'] is not None else '-',
                    round(linha[f'p{prompt_num}_gpu_temp'], 1) if linha[f'p{prompt_num}_gpu_temp'] is not None else '-',
                    linha[f'p{prompt_num}_total_cam'] if linha[f'p{prompt_num}_total_cam'] is not None else '-',
                    linha[f'p{prompt_num}_cam_on'] if linha[f'p{prompt_num}_cam_on'] is not None else '-',
                    linha[f'p{prompt_num}_cam_off'] if linha[f'p{prompt_num}_cam_off'] is not None else '-',
                    linha[f'p{prompt_num}_cam_idle'] if linha[f'p{prompt_num}_cam_idle'] is not None else '-',
                ]
                ws.append(row_data)

            # Ajusta largura das colunas
            ws.column_dimensions['A'].width = 20  # Data/Hora
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                ws.column_dimensions[col].width = 12

            # Centraliza todas as células de dados
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(cabecalhos)):
                for cell in row:
                    cell.alignment = cell_alignment

        # Salva o arquivo em memória
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nome do arquivo com timestamp
        horario_brasil = obter_horario_brasil()
        filename = f"grafana_export_{horario_brasil.strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"Erro ao exportar Excel: {e}")
        import traceback
        traceback.print_exc()
        return f"Erro ao gerar planilha: {str(e)}", 500


def coletar_dados_periodicamente():
    """Função que roda em background coletando dados a cada 10 minutos"""
    print("=" * 70)
    print("COLETA GRAFANA - VERSÃO WEB - 10 MINUTOS")
    print("SALVANDO NO BANCO DE DADOS")
    print("=" * 70 + "\n")

    while True:
        dados = extrair_dados_selenium()

        if dados:
            salvar_em_banco(dados)
        else:
            print("  ❌ Falha na coleta\n")

        print("  Aguardando 10 minutos...\n")
        time.sleep(600)  # 10 minutos


# Inicia a thread de coleta automaticamente quando o módulo é carregado
# Isso garante que funcione tanto com execução direta quanto com Gunicorn
thread_coleta = threading.Thread(target=coletar_dados_periodicamente, daemon=True)
thread_coleta.start()
print("[INFO] Thread de coleta iniciada em background")


if __name__ == "__main__":
    # Quando executado diretamente (não com Gunicorn)
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
